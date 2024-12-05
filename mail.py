import openpyxl
from openpyxl import Workbook
from pymongo import MongoClient
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# Connect to MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['mydatabase']
collection = db['mycollection']

# Open the Excel file
wb = openpyxl.load_workbook('testwb.xlsx')

# Create a new sheet for today's date
today = datetime.today().strftime('%Y-%m-%d')
ws = wb.create_sheet(title=today)

# Write headers to the new sheet
headers = list(collection.find_one().keys())
for i, header in enumerate(headers):
    ws.cell(row=1, column=i+1).value = header

# Write data to the new sheet
data = list(collection.find())
for i, document in enumerate(data):
    for j, value in enumerate(document.values()):
        ws.cell(row=i+2, column=j+1).value = value

# Save the workbook
wb.save('testwb.xlsx')

# Open the log sheet
log_ws = wb['log']

# Write the update time to the log sheet
update_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
log_ws.append([update_time])

# Save the workbook
wb.save('testwb.xlsx')

# Email configuration
smtp_server = 'smtp.gmail.com'
smtp_port = 587
from_email = 'your-email@gmail.com'
password = 'your-password'
to_emails = ['recipient1@example.com', 'recipient2@example.com']

# Create a message
msg = MIMEMultipart()
msg['From'] = from_email
msg['To'] = ', '.join(to_emails)
msg['Subject'] = f'Today\'s data {today}'
msg['Date'] = formatdate(localtime=True)

# Attach the workbook
attachment = open('testwb.xlsx', 'rb')
part = MIMEBase('application', 'octet-stream')
part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', 'attachment; filename= testwb.xlsx')
msg.attach(part)

# Send the email
server = smtplib.SMTP(smtp_server, smtp_port)
server.starttls()
server.login(from_email, password)
server.sendmail(from_email, to_emails, msg.as_string())
server.quit()

print('Email sent successfully!')
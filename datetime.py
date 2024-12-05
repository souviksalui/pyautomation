# This code assumes that you have an Excel file named testwb.xlsx 
# with a sheet named log that records the updates. 
# The code opens the Excel file, creates a new sheet for today's date,
# writes the headers to the new sheet, writes the data to the new sheet, saves the workbook, 
# and appends the update time to the log sheet.

# Note that this code assumes that the mydatabase 
# and mycollection are the names of your MongoDB database and collection, respectively. 
# You may need to modify the code to match your specific setup.


# To run it automatically every day, you'll need to schedule it using a scheduler.

# Here are a few options to schedule the script:

# Cron job (Linux/Mac): You can use the cron command to schedule the script to run daily. Open the terminal and type crontab -e to edit the cron table. Add the following line to schedule the script to run daily at 12:00 AM:
# bash
# CopyInsert in Terminal
# 0 0 * * * python /path/to/your/script.py
# Task Scheduler (Windows): You can use the Task Scheduler to schedule the script to run daily. Open the Task Scheduler and create a new task. Set the trigger to "Daily" and the action to "Start a program" with the Python executable and the script path as arguments.
# Schedule library (Python): You can use the schedule library in Python to schedule the script to run daily. Add the following code to your script:
# python
# CopyInsert
# import schedule
# import time

# def run_script():
#     # Your script code here

# schedule.every().day.at("00:00").do(run_script)  # Run daily at 12:00 AM

# while True:
#     schedule.run_pending()
#     time.sleep(1)
# Cloud scheduling services: You can also use cloud scheduling services like AWS Lambda, Google Cloud Scheduler, or Microsoft Azure Scheduler to schedule your script to run daily.
# Once you've set up the scheduling, your script will run automatically every day at the specified time.



import openpyxl
from openpyxl import Workbook
from pymongo import MongoClient
from datetime import datetime

# Connect to MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['mydatabase']
collection = db['mycollection']

# Open the Excel file
wb = openpyxl.load_workbook('testwb.xlsx')

# Create a new sheet for today's date
today = datetime.today().strftime('%Y-%m-%d')
ws = wb.create_sheet(title=today)

# Write headers to the new sheet
headers = list(collection.find_one().keys())
for i, header in enumerate(headers):
    ws.cell(row=1, column=i+1).value = header

# Write data to the new sheet
data = list(collection.find())
for i, document in enumerate(data):
    for j, value in enumerate(document.values()):
        ws.cell(row=i+2, column=j+1).value = value

# Save the workbook
wb.save('testwb.xlsx')

# Open the log sheet
log_ws = wb['log']

# Write the update time to the log sheet
update_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
log_ws.append([update_time])

# Save the workbook
wb.save('testwb.xlsx')
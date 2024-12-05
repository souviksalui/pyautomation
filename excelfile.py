import openpyxl
from openpyxl import Workbook

def write_to_excel(data_map, file_path, file_name):
    # Create a new Excel file
    wb = Workbook()
    ws = wb.active

    # Extract headers and data from the map
    headers = list(data_map.keys())
    data = list(data_map.values())

    # Write headers to the Excel sheet
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1).value = header

    # Write data to the Excel sheet
    for i, value in enumerate(data):
        ws.cell(row=2, column=i+1).value = value

    # Save the workbook to the file
    wb.save(f"{file_path}/{file_name}.xlsx")

def main():
    # Create a new dictionary to store data
    data_map = {
        "AN_Rate": "Data-1",
        "BN_Rate": "Data-2",
        "CN_Rate": "Data-3"
    }

    # Update the file path
    file_path = "path/to/your/file"
    file_name = "demo-data"

    # Write data to the Excel file
    write_to_excel(data_map, file_path, file_name)

if __name__ == "__main__":
    main()
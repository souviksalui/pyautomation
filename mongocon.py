import openpyxl
from openpyxl import Workbook
from pymongo import MongoClient

# Connect to MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['mydatabase']
collection = db['mycollection']

# Retrieve data from MongoDB collection
data = list(collection.find())

# Create a new Excel file
wb = Workbook()
ws = wb.active

# Write headers to the Excel sheet
headers = list(data[0].keys())
for i, header in enumerate(headers):
    ws.cell(row=1, column=i+1).value = header

# Write data to the Excel sheet
for i, document in enumerate(data):
    for j, value in enumerate(document.values()):
        ws.cell(row=i+2, column=j+1).value = value

# Save the workbook to the file
wb.save('output.xlsx')